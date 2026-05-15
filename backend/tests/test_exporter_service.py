import io
from types import SimpleNamespace

from openpyxl import load_workbook

from app.services.exporter import (
    build_export_rows,
    export_clean_records_to_csv,
    export_clean_records_to_xlsx,
    export_rows_to_csv,
    export_rows_to_xlsx,
)



def test_build_export_rows_maps_clean_records_to_template_columns() -> None:
    clean_records = [
        SimpleNamespace(
            clean_payload={
                "name": "Example University",
                "website": "https://example.edu",
            }
        ),
        SimpleNamespace(
            clean_payload={
                "name": "Sample Institute",
                "website": "   ",
            }
        ),
    ]
    template_columns = [
        {"name": "slug", "order": 2},
        {"name": "name", "order": 1},
        {"name": "website", "order": 3},
        {"name": "sponsored", "order": 4},
    ]

    rows = build_export_rows(
        clean_records,
        template_columns=template_columns,
        defaults={"website": "https://fallback.example.edu"},
    )

    assert rows == [
        {
            "name": "Example University",
            "slug": "example-university",
            "website": "https://example.edu",
            "sponsored": False,
        },
        {
            "name": "Sample Institute",
            "slug": "sample-institute",
            "website": "https://fallback.example.edu",
            "sponsored": False,
        },
    ]



def test_export_clean_records_to_csv_maps_domain_records_before_writing_csv() -> None:
    clean_records = [
        SimpleNamespace(
            clean_payload={
                "name": "Example University",
                "website": "https://example.edu",
            }
        ),
        SimpleNamespace(
            clean_payload={
                "name": "Sample Institute",
            }
        ),
    ]
    template_columns = [
        {"name": "name", "order": 1},
        {"name": "slug", "order": 2},
        {"name": "website", "order": 3},
    ]

    content = export_clean_records_to_csv(
        clean_records,
        template_columns=template_columns,
        defaults={"website": "https://fallback.example.edu"},
    )

    assert content.decode("utf-8") == (
        "name,slug,website\r\n"
        "Example University,example-university,https://example.edu\r\n"
        "Sample Institute,sample-institute,https://fallback.example.edu\r\n"
    )



def test_export_rows_to_csv_preserves_template_header_order() -> None:
    rows = [
        {
            "name": "Example University",
            "website": "https://example.edu",
            "id": None,
        }
    ]
    template_columns = [
        {"name": "id", "order": 1},
        {"name": "name", "order": 2},
        {"name": "website", "order": 3},
    ]

    content = export_rows_to_csv(rows, template_columns=template_columns)

    assert content.decode("utf-8") == "id,name,website\r\n,Example University,https://example.edu\r\n"



def test_export_rows_to_csv_exports_multiple_rows_in_template_order() -> None:
    rows = [
        {
            "website": "https://example.edu",
            "name": "Example University",
            "slug": "example-university",
        },
        {
            "website": "https://sample.edu",
            "name": "Sample Institute",
            "slug": "sample-institute",
        },
    ]
    template_columns = [
        {"name": "name", "order": 2},
        {"name": "slug", "order": 3},
        {"name": "website", "order": 1},
    ]

    content = export_rows_to_csv(rows, template_columns=template_columns)

    assert content.decode("utf-8") == (
        "website,name,slug\r\n"
        "https://example.edu,Example University,example-university\r\n"
        "https://sample.edu,Sample Institute,sample-institute\r\n"
    )



def test_export_rows_to_csv_ignores_extra_fields_and_leaves_missing_cells_blank() -> None:
    rows = [
        {
            "name": "Example University",
            "internal_note": "ignore me",
        }
    ]
    template_columns = [
        {"name": "name", "order": 1},
        {"name": "website", "order": 2},
    ]

    content = export_rows_to_csv(rows, template_columns=template_columns)

    assert content.decode("utf-8") == "name,website\r\nExample University,\r\n"



def test_export_rows_to_xlsx_preserves_template_header_order() -> None:
    rows = [
        {
            "website": "https://example.edu",
            "name": "Example University",
            "slug": "example-university",
        }
    ]
    template_columns = [
        {"name": "name", "order": 2},
        {"name": "slug", "order": 3},
        {"name": "website", "order": 1},
    ]

    content = export_rows_to_xlsx(rows, template_columns=template_columns)

    workbook = load_workbook(io.BytesIO(content))
    worksheet = workbook.active
    assert list(worksheet.iter_rows(values_only=True)) == [
        ("website", "name", "slug"),
        ("https://example.edu", "Example University", "example-university"),
    ]



def test_export_clean_records_to_xlsx_maps_domain_records_before_writing_workbook() -> None:
    clean_records = [
        SimpleNamespace(
            clean_payload={
                "name": "Example University",
                "website": "https://example.edu",
            }
        ),
        SimpleNamespace(
            clean_payload={
                "name": "Sample Institute",
            }
        ),
    ]
    template_columns = [
        {"name": "name", "order": 1},
        {"name": "slug", "order": 2},
        {"name": "website", "order": 3},
    ]

    content = export_clean_records_to_xlsx(
        clean_records,
        template_columns=template_columns,
        defaults={"website": "https://fallback.example.edu"},
    )

    workbook = load_workbook(io.BytesIO(content))
    worksheet = workbook.active
    assert list(worksheet.iter_rows(values_only=True)) == [
        ("name", "slug", "website"),
        ("Example University", "example-university", "https://example.edu"),
        ("Sample Institute", "sample-institute", "https://fallback.example.edu"),
    ]



def test_export_rows_to_xlsx_ignores_extra_fields_and_leaves_missing_cells_blank() -> None:
    rows = [
        {
            "name": "Example University",
            "internal_note": "ignore me",
        }
    ]
    template_columns = [
        {"name": "name", "order": 1},
        {"name": "website", "order": 2},
    ]

    content = export_rows_to_xlsx(rows, template_columns=template_columns)

    workbook = load_workbook(io.BytesIO(content))
    worksheet = workbook.active
    assert list(worksheet.iter_rows(values_only=True)) == [
        ("name", "website"),
        ("Example University", None),
    ]



def test_export_rows_to_xlsx_returns_workbook_bytes() -> None:
    content = export_rows_to_xlsx([], template_columns=[{"name": "name", "order": 1}])

    assert isinstance(content, bytes)
    assert content[:2] == b"PK"
